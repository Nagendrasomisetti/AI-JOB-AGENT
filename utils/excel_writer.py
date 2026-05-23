import os
import datetime
import pandas as pd
import logging
from typing import List, Dict, Any

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

logger = logging.getLogger("AIJobAgent.ExcelWriter")

def save_jobs_to_excel(jobs: List[Dict[str, Any]]) -> str:
    """
    Saves and formats a list of jobs into a highly polished, styled Excel sheet.
    Incorporates dynamic column resizing, themed headers, and functional hyperlinks.
    
    Args:
        jobs (List[Dict]): List of processed job listings.
        
    Returns:
        str: Filename of the generated spreadsheet.
    """
    if not jobs:
        logger.info("ExcelWriter received empty job list; skipping write.")
        return ""

    df = pd.DataFrame(jobs)

    # 1. Enforce strict column order for visual consistency
    ordered_columns = [
        "title",
        "company",
        "location",
        "salary",
        "type",
        "experience",
        "skills",
        "link",
        "source",
        "score"
    ]
    
    # Filter columns to only keep what is available in the dataframe
    cols_to_use = [col for col in ordered_columns if col in df.columns]
    df = df[cols_to_use]

    # Clean up column headers for a premium presentation
    header_mapping = {col: col.replace("_", " ").title() for col in cols_to_use}
    # Keep link as "Apply Link"
    if "link" in header_mapping:
        header_mapping["link"] = "Apply Link"
    
    df = df.rename(columns=header_mapping)

    # 2. Establish filename with robust timestamping
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"jobs_{timestamp}.xlsx"
    
    logger.info(f"Writing {len(jobs)} jobs to raw Excel structure: {filename}...")
    
    try:
        # Write to spreadsheet using pandas engine
        df.to_excel(filename, index=False)
        
        # 3. Apply Premium Openpyxl styling if the library is available
        if OPENPYXL_AVAILABLE:
            logger.info("Applying premium visual styles to Excel sheet...")
            _style_workbook(filename)
            logger.info(f"Excel file styled and saved: {filename}")
        else:
            logger.warning("openpyxl is not installed. Saved basic, unstyled Excel spreadsheet.")
            
        return filename

    except Exception as e:
        logger.error(f"Failed to generate Excel report: {e}")
        return ""


def _style_workbook(filepath: str) -> None:
    """
    Injects custom styles, borders, headers, and hyperlinks into the generated Excel sheet.
    """
    wb = load_workbook(filepath)
    ws = wb.active
    
    # Color palette tokens
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid") # Deep Charcoal Blue
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    score_font = Font(name="Calibri", size=10, bold=True)
    link_font = Font(name="Calibri", size=10, color="1B4F72", underline="single") # Professional dark blue link
    
    # Alignments
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    right_align = Alignment(horizontal="right", vertical="center")
    
    # Borders
    thin_border_side = Side(border_style="thin", color="BDC3C7")
    cell_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )
    
    # Get column header row (Row 1)
    ws.row_dimensions[1].height = 28
    
    headers = [cell.value for cell in ws[1]]
    link_col_idx = -1
    score_col_idx = -1
    
    for idx, header in enumerate(headers):
        if header == "Apply Link":
            link_col_idx = idx + 1
        elif header == "Score":
            score_col_idx = idx + 1

    # Style Header Row
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = cell_border

    # Style Data Rows
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = 20
        
        # Apply standard alignments and thin borders to every cell
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = data_font
            cell.border = cell_border
            
            # Default text alignment
            if col_idx in [3, 4, 5, 6, 9]: # Location, Salary, Type, Exp, Source
                cell.alignment = center_align
            else:
                cell.alignment = left_align

            # Special column treatments
            if col_idx == score_col_idx:
                cell.font = score_font
                cell.alignment = center_align
            
            elif col_idx == link_col_idx:
                # Convert raw URLs into sleek clickable =HYPERLINK formulas
                url = cell.value
                if url and str(url).startswith("http"):
                    cell.value = f'=HYPERLINK("{url}", "Apply ↗")'
                    cell.font = link_font
                    cell.alignment = center_align

    # 4. Auto-fit columns dynamically based on text lengths
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        
        for cell in col:
            val_str = str(cell.value or "")
            # Skip checking the raw hyperlink formulas which skew length calculations
            if "=HYPERLINK" in val_str:
                max_len = max(max_len, 8)
            else:
                max_len = max(max_len, len(val_str))
                
        # Set a comfortable padding (max_len + 3), clamped between 10 and 45 characters
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 10), 45)

    # Save the modified sheet back
    wb.save(filepath)
    wb.close()